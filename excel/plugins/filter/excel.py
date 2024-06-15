from openpyxl import *

class FilterModule(object):
    def filters(self):
        return {
            'read_excel': self.ReadExcel,
            'read_excel_range': self.ReadExcelRange,
            'read_excel_header': self.ReadExcelHeader,
            'read_excel_header_range': self.ReadExcelHeaderRange,
        }

    def ReadExcel(self, path, *args):
        workbook = load_workbook(filename=path, data_only=True)
        sheet = workbook[args[0]] if len(args) != 0 else workbook[workbook.sheetnames[0]]
        
        values = []
        for r in range (1, sheet.max_row + 1):
            temp = []
            for c in range(1, sheet.max_column + 1):
                temp.append(sheet.cell(row=r,column=c).value)
            values.append(temp)
        return values

    def ReadExcelRange(self, path, startrow, startcolumn, endrow, endcolumn, *args):
        workbook = load_workbook(filename=path, data_only=True)
        sheet = workbook[args[0]] if len(args) != 0 else workbook[workbook.sheetnames[0]]
        
        if endrow == 0:
            endrow = sheet.max_row
        if endcolumn == 0:
            endcolumn = sheet.max_column

        values = []
        for r in range (startrow, endrow + 1):
            temp = []
            for c in range(startcolumn, endcolumn + 1):
                temp.append(sheet.cell(row=r,column=c).value)
            values.append(temp)
        return values

    def ReadExcelHeader(self, path, *args):
        workbook = load_workbook(filename=path, data_only=True)
        sheet = workbook[args[0]] if len(args) != 0 else workbook[workbook.sheetnames[0]]
        
        values = []
        header = []
        for c in range(1, sheet.max_column + 1):
            header.append(sheet.cell(1,column=c).value)
        for r in range (2, sheet.max_row + 1):
            temp = {}
            for c in range(1, sheet.max_column + 1):
                temp[header[c - 1]] = sheet.cell(row=r,column=c).value
            values.append(temp)
        return values

    def ReadExcelHeaderRange(self, path, startrow, startcolumn, endrow, endcolumn, *args):
        workbook = load_workbook(filename=path, data_only=True)
        sheet = workbook[args[0]] if len(args) != 0 else workbook[workbook.sheetnames[0]]

        if endrow == 0:
            endrow = sheet.max_row
        if endcolumn == 0:
            endcolumn = sheet.max_column

        values = []
        header = []
        for c in range(startcolumn, endcolumn + 1):
            header.append(sheet.cell(row=startrow,column=c).value)
        for r in range (startrow + 1, endrow + 1):
            temp = {}
            for c in range(startcolumn, endcolumn + 1):
                temp[header[c - 1]] = sheet.cell(row=r,column=c).value
            values.append(temp)
        return values
