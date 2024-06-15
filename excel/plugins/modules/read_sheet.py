#!/usr/bin/python
# -*- coding: utf-8 -*-

# Copyright: (c) 2024, Yuichi Mizutani (yumizu11) <iyumizu@hotmail.com>
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = '''
---
module: read_sheet

short_description: This module read Excel file and return cell values

version_added: "0.1"

description:
    - "This module read a specified sheet in an Excel file (.xslx), and return rows of cell values, or map if the 1st row is header line. You can specify the range of cells to read."

options:
    path:
        description:
            - Path to the Excel file you want to read
        required: true
    sheet:
        description:
            - Name of sheet you want to read. The default is the 1st sheet in the Excel file
        required: false
    range:
        description:
            - Range of cell to read. This range is Excel style range, like 'A2:E5'. 'B2:' means starting with B2 to the end.
        required: false
    startrow:
        description:
            - Row number to start reading from. The default is 1. If range is specified, this param will be ignored.
        required: false
    startcolumn:
        description:
            - Column number to start reading from. The default is 1, meaning A.  If range is specified, this param will be ignored.
        required: false
    endrow:
        description:
            - Row number to finish reading. The last row in the sheet if not specified.  If range is specified, this param will be ignored.
        required: false
    endcolumn:
        description:
            - Column number to finish reading. The last column in the sheet if not specified.  If range is specified, this param will be ignored.
        required: false
    header:
        description:
            - Specify 'true', if the first row is the header line. The default is 'false'
        required: false

author:
    - Yuichi Mizutani (@yumizu11)
'''

EXAMPLES = '''
# Read a whole rows of the 1st sheet
- name: Read whole rows
  yumizu11.excel.read_sheet:
    path: mysheet.xlsx
  register: sheet_rows

# Read the specified sheet
- name: Read whole rows of mysheet
  yumizu11.excel.read_sheet:
    path: mysheet.xlsx
    sheet: mysheet
  register: mysheet_rows

# Read specified range of cells
- name: Read B2 to D5
  yumizu11.excel.read_sheet:
    path: mysheet.xlsx
    startrow: 2
    endrow: 5
    startcolumn: 2
    endcolumn: 5
  register: sheet_b2_d5

# Read specified range of cells
- name: Read sheet with header in the 2nd row
  yumizu11.excel.read_sheet:
    path: mysheet.xlsx
    startrow: 2
    header: true
  register: sheet_rows
'''

RETURN = '''
rows:
    description: List of rows read in the sheet. Each row has list of cell values. You can access cell value directly with rows[row][column].
    type: list
    returned: always
path:
    description: The path to the Excel file the module read
    type: str
    returned: always
sheet:
    description: Sheet name the module read
    type: str
    returned: always
'''

from ansible.module_utils.basic import *
from openpyxl import *
import re

def main():
    module = AnsibleModule(argument_spec = dict(
             path = dict(required=True, type='str'),
             sheet = dict(required=False, type='str', default=''),
             startrow = dict(required=False, type='int', default=1),
             startcolumn = dict(required=False, type='int', default=1),
             endrow = dict(required=False, type='int', default=0),
             endcolumn = dict(required=False, type='int', default=0),
             range = dict(required=False, type='str', default=''),
             header = dict(required=False, type='bool', default=False),
             ),
             add_file_common_args=True)

    result = {}
    try:
        # check file existance, and open it
        result["path"] = module.params["path"]
        if os.path.isfile(module.params["path"]) == False:
            module.fail_json(msg="Excel file '" + module.params["path"] + "' does not exist.")
            return 1

        # check sheet
        workbook = load_workbook(filename=module.params["path"], data_only=True)
        sheetname = module.params["sheet"]
        if sheetname == "":
            sheetname = workbook.sheetnames[0]
        if sheetname not in workbook.sheetnames:
            module.fail_json(msg="Sheet '" + sheetname + "' does not exist.")
            return 1
        sheet = workbook[sheetname]
        result["sheet"] = sheetname

        startrow = 0
        startcolumn = 0
        endrow = 0
        endcolumn = 0

        # check range
        cellrange = module.params["range"]
        if cellrange != '':
            # if ":" not in cellrange:
            #     module.fail_json("range parameter must contain ':' character.")
            #     return 1

            reresult = re.search(r'([A-Z]*)(\d*):([A-Z]*)(\d*)', cellrange)

            startcolumn_s = reresult.group(1)
            if startcolumn_s:
                if len(startcolumn_s) == 3:
                    startcolumn = (ord(startcolumn_s[0]) - 0x40) * 26 * 26 + (ord(startcolumn_s[1]) - 0x40) * 26 + (ord(startcolumn_s[2]) - 0x40)
                if len(startcolumn_s) == 2:
                    startcolumn = (ord(startcolumn_s[0]) - 0x40) * 26 + (ord(startcolumn_s[1]) - 0x40)
                if len(startcolumn_s) == 1:
                    startcolumn = ord(startcolumn_s[0]) - 0x40
            else:
                startcolumn = 1

            startrow_w = reresult.group(2)
            if startrow_w:
                startrow = int(startrow_w)
            else:
                startrow = 1

            endcolumn_w = reresult.group(3)
            if endcolumn_w:
                endcolumn_s = endcolumn_w
                if len(endcolumn_s) == 3:
                    endcolumn = (ord(endcolumn_s[0]) - 0x40) * 26 * 26 + (ord(endcolumn_s[1]) - 0x40) * 26 + (ord(endcolumn_s[2]) - 0x40)
                if len(endcolumn_s) == 2:
                    endcolumn = (ord(endcolumn_s[0]) - 0x40) * 26 + (ord(endcolumn_s[1]) - 0x40)
                if len(endcolumn_s) == 1:
                    endcolumn = ord(endcolumn_s[0]) - 0x40
            else:
                if ":" not in cellrange:
                    endcolumn = startcolumn
                else:
                    endcolumn = sheet.max_column

            endrow_w = reresult.group(4)
            if endrow_w:
                endrow = int(endrow_w)
            else:
                if ":" not in cellrange:
                    endrow = startrow
                else:
                    endrow = sheet.max_row
        else:
            startrow = module.params["startrow"]
            startcolumn = module.params["startcolumn"]
            endrow = module.params["endrow"] if module.params["endrow"] != 0 else sheet.max_row
            endcolumn = module.params["endcolumn"] if module.params["endcolumn"] != 0 else sheet.max_column

        result["startrow"] = startrow
        result["startcolumn"] = startcolumn
        result["endrow"] = endrow
        result["endcolumn"] = endcolumn

        if startrow < 1:
            module.fail_json(msg="startrow is smaller than 1")
            return 1
        if startcolumn < 1:
            module.fail_json(msg="startcolumn is smaller than 1")
            return 1
        if endrow < startrow:
            module.fail_json(msg="endrow is smaller than startrow")
            return 1
        if endcolumn < startcolumn:
            module.fail_json(msg="endcolumn is smaller than startcolumn")
            return 1

        values = []
        if module.params["header"]:
            # read with header
            header = []
            for c in range(startcolumn, endcolumn + 1):
                h = sheet.cell(row=startrow,column=c).value
                if h in header:
                    pf = 2
                    h2 = h
                    while h2 in header:
                        h2 = h + "_" + str(pf)
                        pf = pf + 1
                    h = h2
                header.append(h)
            for r in range (startrow + 1, endrow):
                temp = {}
                for c in range(startcolumn, endcolumn):
                    temp[header[c - 1]] = sheet.cell(row=r,column=c).value
                values.append(temp)
        else:
            # read without header
            for r in range(startrow, endrow + 1):
                temp = []
                for c in range(startcolumn, endcolumn + 1):
                    temp.append(sheet.cell(row=r,column=c).value)
                values.append(temp)
        result["rows"] = values
        module.exit_json(**result)
    except IOError:
        module.fail_json(msg="Error on loading excel file " + module.params["path"])
        return 1

    return 0

main()
