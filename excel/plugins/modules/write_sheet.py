#!/usr/bin/python
# -*- coding: utf-8 -*-

# Copyright: (c) 2024, Yuichi Mizutani (yumizu11) <iyumizu@hotmail.com>
# GNU General Public License v3.0+ (see COPYING or https://www.gnu.org/licenses/gpl-3.0.txt)

DOCUMENTATION = '''
---
module: write_sheet

short_description: This module write values to the specified sheet in the specified Excel file

version_added: "0.1"

description:
    - "This module writes a list of list or a list of map to the specified sheet in the specfied Excel file."
    - "If the data is a list of map, keys will be written on the top row, and data follow."
    - "Rows can be inserted by setting param 'insert' to true, otherwise cells will be overwritten."
    - "Excel file will be created, if not exist."

options:
    path:
        description:
            - Path to the Excel file you want to write data
        required: true
    sheet:
        description:
            - Name of sheet you want to write data. 1st sheet in the Excel file will be used if not specified
        required: false
    startrow:
        description:
            - Row number to start writing data from. The default is 1.
        required: false
    startcolumn:
        description:
            - Column number to start writing data from. The default is 1, meaning A column.
        required: false
    data:
        description:
            - list of values to write to the sheet. This can be a list of list or list of map. If it is a list of map, the top row will be a header row.
        required: true
    insert:
        description:
            - If true, empty rows will be inserted and then data will be written on the inserted rows.
        required: false

extends_documentation_fragment:

author:
    - Yuichi Mizutani (@yumizu11)
'''

EXAMPLES = '''
# Write 2D array to the first sheet of mysheet.xlsx
- name: Write whole data
  vars:
    mydata:
      - [1,2,3]
      - [4,5,6]
  yumizu11.excel.write_sheet:
    path: mysheet.xlsx
    data: "{{ mydata }}"

# Write a list of map to sheet 'sheetA'
- name: Write whole data
  vars:
    mydata:
      - firstname: Taro
        lastname: Yamada
      - firstname: Hanako
        lastname: Sato
  yumizu11.excel.write_sheet:
    path: mysheet.xlsx
    sheet: sheetA
    data: "{{ mydata }}"

# Write 2D array to cell C3~ on the first sheet
- name: Write whole data
  vars:
    mydata:
      - [1,2,3]
      - [4,5,6]
  yumizu11.excel.write_sheet:
    path: mysheet.xlsx
    data: "{{ mydata }}"
    startrow: 3
    startcolumn: 3

# Insert a list of map on the cell C3
- name: Write whole data
  vars:
    mydata:
      - firstname: Taro
        lastname: Yamada
      - firstname: Hanako
        lastname: Sato
  yumizu11.excel.write_sheet:
    path: mysheet.xlsx
    sheet: sheetA
    data: "{{ mydata }}"
    startrow: 3
    startcolumn: 3
    insert: True
'''

RETURN = '''
path:
    description: The path to the Excel file the module wrote
    type: str
    returned: always
sheet:
    description: Sheet name the module wrote
    type: str
    returned: always
'''

from ansible.module_utils.basic import *
from openpyxl import *

def row_column_to_cell_name(row, column):
    cell = ""
    if column > 26 * 26:
        cell = chr(0x40 + ((column - 1) // (26 * 26 -1)) + 1)
    if column > 26:
        cell += chr(0x40 + ((column - 1) // 25) + 1)
    cell += chr(0x40 + ((column - 1) % 26) + 1) + str(row)
    return cell

def main():
    module = AnsibleModule(argument_spec = dict(
             path = dict(required=True, type='str'),
             sheet = dict(required=False, type='str', default=''),
             startrow = dict(required=False, type='int', default=1),
             startcolumn = dict(required=False, type='int', default=1),
             data = dict(required=True, type='list'),
             insert = dict(required=False, type='bool', default=False),
             ),
             add_file_common_args=True)

    result = {}
    changed = False
    try:
        result["path"] = module.params["path"]
        if os.path.isfile(module.params["path"]) == False:
            create = True
            changed = True
            workbook = Workbook()
        else:
            workbook = load_workbook(filename=module.params["path"])

        sheet = workbook.active
        sheetname = module.params["sheet"]
        if sheetname != '':
            if sheetname not in workbook.sheetnames:
                workbook.create_sheet(sheetname)
                if create:
                    workbook.remove(workbook.active)
            sheet = workbook[sheetname]
        result["sheet"] = sheetname

        data = module.params["data"]

        if len(data) > 0:
            if type(data[0]) == list:
                if module.params["insert"]:
                    # check idempotence before inserting rows
                    need_change = False
                    crow = module.params["startrow"]
                    for arow in data:
                        column = module.params["startcolumn"]
                        for value in arow:
                            if sheet[row_column_to_cell_name(crow, column)].value != value:
                                need_change = True
                            column = column + 1
                        crow = crow + 1
                    if need_change == False:
                        result["changed"] = False
                        module.exit_json(**result)
                        return 0
                    # insert rows
                    sheet.insert_rows(row, len(data))
                # write data to sheet
                for rowdata in data:
                    row = module.params["startrow"]
                    column = module.params["startcolumn"]
                    for value in rowdata:
                        if sheet[row_column_to_cell_name(row, column)].value != value:
                            sheet[row_column_to_cell_name(row, column)] = value
                            changed = True
                        column = column + 1
                    row = row + 1
            else:
                # gather header
                header = {}
                column = module.params["startcolumn"]
                for rowdata in data:
                    for k in rowdata.keys():
                        if k not in header.keys():
                            header[k] = column
                            column = column + 1

                if module.params["insert"]:
                    # check idempotence before inserting rows
                    need_change = False
                    crow = module.params["startrow"]
                    for k in header.keys():
                        if sheet[row_column_to_cell_name(crow, header[k])].value != k:
                            need_change = True
                    crow = crow + 1
                    for rowdata in data:
                        for k in rowdata.keys():
                            if sheet[row_column_to_cell_name(crow, header[k])].value != rowdata[k]:
                                need_change = True
                        crow = crow + 1
                    if need_change == False:
                        result["changed"] = False
                        module.exit_json(**result)
                        return 0
                    # insert rows
                    sheet.insert_rows(module.params["startrow"], len(data) + 1)

                # write data to sheet
                row = module.params["startrow"]
                for k in header.keys():
                    if create or sheet[row_column_to_cell_name(row, header[k])].value != k:
                        sheet[row_column_to_cell_name(row, header[k])] = k
                        changed = True
                row = row + 1
                for rowdata in data:
                    for k in rowdata.keys():
                        if create or sheet[row_column_to_cell_name(row, header[k])].value != rowdata[k]:
                            sheet[row_column_to_cell_name(row, header[k])] = rowdata[k]
                            changed = True
                    row = row + 1

        if changed:
            workbook.save(module.params["path"])
        result["changed"] = changed
        module.exit_json(**result)
    except IOError:
        module.fail_json(msg="Error on writing excel file")
        return 1

    return 0

main()
